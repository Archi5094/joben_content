import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

# --- Mappings ---
territory_region_map = {
    'T1': 'QC', 'T2': 'QC', 'T3': 'QC',
    'T4': 'ON', 'T5': 'ON', 'T6': 'ON',
    'T7': 'BC', 'T8': 'PR', 'T9': 'AR',
    'Territory T1': 'QC', 'Territory T2': 'QC', 'Territory T3': 'QC',
    'Territory T4': 'ON', 'Territory T5': 'ON', 'Territory T6': 'ON',
    'Territory T7': 'BC', 'Territory T8': 'PR', 'Territory T9': 'AR'
}

model_map = {
    'CO45': 'Outlander', 'COEV': 'Outlander PHEV',
    'CE45': 'Eclipse Cross', 'CS45': 'RVR', 'CG44': 'Mirage'
}

origin_map = {
    'ClickShop': 'ClickShop',
    'MMSCAN Brand Website': 'CWS',
    'Autotrader Storefront': 'AutoTrader',
    'Meta': 'Meta'
}

model_row_map = {'Outlander': 69, 'Outlander PHEV': 70, 'Eclipse Cross': 71, 'RVR': 72, 'Mirage': 73}
origin_row_map = {'ClickShop': 44, 'CWS': 45, 'AutoTrader': 46, 'Meta': 47}
footfall_row_map = {'Outlander': 62, 'Outlander PHEV': 63, 'Eclipse Cross': 64, 'RVR': 65, 'Mirage': 66}
region_col_map = {'BC': 'L', 'PR': 'M', 'ON': 'N', 'QC': 'O', 'AR': 'P'}

def get_sheet_names(uploaded_file):
    if uploaded_file is None:
        return []
    try:
        uploaded_file.seek(0)
        xls = pd.ExcelFile(uploaded_file)
        uploaded_file.seek(0)
        return xls.sheet_names
    except Exception as e:
        st.error(f"Error reading sheets: {e}")
        return []

st.title("📊 Weekly Regional Report Generator")

# File Uploaders
report_file = st.file_uploader("Upload Report File", type=["xlsx"])
report_sheet = st.selectbox("Select Report Sheet", get_sheet_names(report_file)) if report_file else None

sales_file = st.file_uploader("Upload Retail Sales File", type=["xlsx"])
sales_sheet = st.selectbox("Select Retail Sales Sheet", get_sheet_names(sales_file)) if sales_file else None

leads_file = st.file_uploader("Upload Corporate Leads File", type=["xlsx"])
leads_sheet = st.selectbox("Select Corporate Leads Sheet", get_sheet_names(leads_file)) if leads_file else None

footfall_file = st.file_uploader("Upload Footfall File", type=["xlsx"])
footfall_sheet = st.selectbox("Select Footfall Sheet", get_sheet_names(footfall_file)) if footfall_file else None

if st.button("✅ Generate Updated Report"):
    if not all([report_file, report_sheet, sales_file, sales_sheet, leads_file, leads_sheet, footfall_file, footfall_sheet]):
        st.error("Please upload all required files and select sheets.")
    else:
        try:
            # Load Report Workbook
            report_file.seek(0)
            report_bytes = BytesIO(report_file.read())
            report_bytes.seek(0)
            wb = load_workbook(filename=report_bytes, data_only=True)
            ws = wb[report_sheet]

            # Load Footfall Data
            footfall_file.seek(0)
            ff_df = pd.read_excel(footfall_file, sheet_name=footfall_sheet)
            #st.write("Footfall Data Sample:", ff_df.head())
            ff_df['Region'] = ff_df['Region'].map(territory_region_map)
            ff_df.dropna(subset=['Region', 'Model', 'Traffic'], inplace=True)
            ff_pivot = pd.pivot_table(ff_df, index='Model', columns='Region', values='Traffic', aggfunc='sum', fill_value=0)
            st.write("Footfall Pivot Table:", ff_pivot)

            for model, row in footfall_row_map.items():
                if model in ff_pivot.index:
                    for region, col in region_col_map.items():
                        val = int(ff_pivot.loc[model].get(region, 0))
                        #st.write(f"Writing Footfall: ws[{col}{row}] = {val}")
                        ws[f"{col}{row}"] = val

            # Load Retail Sales Data
            sales_file.seek(0)
            rs_df = pd.read_excel(sales_file, sheet_name=sales_sheet)
           # st.write("Retail Sales Data Sample:", rs_df.head())
            rs_df = rs_df[rs_df['Model Code'].isin(model_map)]
            rs_df['Model'] = rs_df['Model Code'].map(model_map)
            rs_df['Region'] = rs_df['Terr.'].map(territory_region_map)
            rs_df.dropna(subset=['Model', 'Region'], inplace=True)
            rs_pivot = pd.pivot_table(rs_df, index='Model', columns='Region', values='Retail Count', aggfunc='sum', fill_value=0)
            st.write("Retail Sales Pivot Table:", rs_pivot)

            for model, row in model_row_map.items():
                if model in rs_pivot.index:
                    for region, col in region_col_map.items():
                        val = int(rs_pivot.loc[model].get(region, 0))
                        #st.write(f"Writing Retail Sales: ws[{col}{row}] = {val}")
                        ws[f"{col}{row}"] = val

            # Load Corporate Leads Data
            leads_file.seek(0)
            cl_df = pd.read_excel(leads_file, sheet_name=leads_sheet)
            #st.write("Corporate Leads Data Sample:", cl_df.head())
            cl_df = cl_df[cl_df['Origin'].isin(origin_map)]
            cl_df['Origin'] = cl_df['Origin'].map(origin_map)
            cl_df['Region'] = cl_df['Territory'].map(territory_region_map)
            cl_df.dropna(subset=['Origin', 'Region'], inplace=True)
            cl_pivot = pd.pivot_table(cl_df, index='Origin', columns='Region', values='Province', aggfunc='count', fill_value=0)
            st.write("Corporate Leads Pivot Table:", cl_pivot)

            for origin, row in origin_row_map.items():
                if origin in cl_pivot.index:
                    for region, col in region_col_map.items():
                        val = int(cl_pivot.loc[origin].get(region, 0))
                       # st.write(f"Writing Corporate Leads: ws[{col}{row}] = {val}")
                        ws[f"{col}{row}"] = val

            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.success("Report generated successfully!")
            st.download_button(
                label="Download Updated Report",
                data=output,
                file_name="Updated_Weekly_Regional_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Error processing files: {e}")
